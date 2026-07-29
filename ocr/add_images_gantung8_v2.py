from pathlib import Path
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

BASE = Path('D:/PROJECT/alfamind/ocr')
IMAGE = BASE / 'Gantung8.jpeg'
WORKBOOK = BASE / 'hasil_gantung8.xlsx'
THUMBS = BASE / 'gantung8_product_images'
THUMBS.mkdir(exist_ok=True)

with Image.open(IMAGE) as source:
    w, h = source.size
    xs = [0, 240, 480, w]
    ys = [140, 410, 670, 940, 1205]
    for i in range(12):
        r, c = divmod(i, 3)
        source.crop((xs[c] + 8, ys[r] + 8, xs[c + 1] - 8, ys[r + 1] - 8)).save(THUMBS / f'product_{i + 1:02d}.png')

wb = openpyxl.load_workbook(WORKBOOK)
ws = wb['Semua Produk']
rows = []
for source_index, values in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True), start=1):
    name = str(values[1] or '')
    parts = [x.strip() for x in name.split(',') if x.strip()] or [name]
    rows.extend((values, part, source_index) for part in parts)

for image in list(ws._images):
    ws._images.remove(image)
ws.delete_rows(4, max(0, ws.max_row - 3))

for row_number, (values, name_part, source_index) in enumerate(rows, start=4):
    for col, value in enumerate(values, start=1):
        ws.cell(row=row_number, column=col).value = name_part if col == 2 else ('' if col == 1 else value)
    ws.row_dimensions[row_number].height = 75
    img = XLImage(str(THUMBS / f'product_{source_index:02d}.png'))
    img.width = 85
    img.height = 85
    ws.add_image(img, f'A{row_number}')

ws['A1'] = f'KATALOG DATA PRODUK & PROMO ALFAMIND ({len(rows)} ITEM)'
wb.save(WORKBOOK)
print(f'Updated {len(rows)} rows with product images')
