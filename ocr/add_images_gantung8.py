from pathlib import Path
from copy import copy

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image


BASE = Path("D:/PROJECT/alfamind/ocr")
IMAGE = BASE / "Gantung8.jpeg"
WORKBOOK = BASE / "hasil_gantung8.xlsx"
THUMBS = BASE / "gantung8_product_images"

# Gantung8 is a 3-column x 4-row product grid below the banner.
with Image.open(IMAGE) as source:
    width, height = source.size
    x_edges = [0, 240, 480, width]
    y_edges = [140, 410, 670, 940, 1205]
    THUMBS.mkdir(exist_ok=True)
    for idx in range(12):
        row, col = divmod(idx, 3)
        x1, x2 = x_edges[col], x_edges[col + 1]
        y1, y2 = y_edges[row], y_edges[row + 1]
        # Keep the product and its label/price while trimming grid borders.
        crop = source.crop((x1 + 8, y1 + 8, x2 - 8, y2 - 8))
        crop.save(THUMBS / f"product_{idx + 1:02d}.png")

wb = openpyxl.load_workbook(WORKBOOK)
ws = wb["Semua Produk"]

# Split comma-delimited product names into separate products. Preserve all
# other specifications and the same product image for each duplicated row.
rows = []
for values in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    name = str(values[1] or "")
    parts = [part.strip() for part in name.split(",") if part.strip()]
    rows.extend([(values, part) for part in (parts or [name])])

# Remove existing data rows and drawings, then rebuild the data rows.
for image in list(ws._images):
    ws._images.remove(image)
if ws.max_row >= 4:
    ws.delete_rows(4, ws.max_row - 3)

for row_number, (values, name_part) in enumerate(rows, start=4):
    output = list(values)
    output[1] = name_part
    output[0] = ""
    for col, value in enumerate(output, start=1):
        ws.cell(row=row_number, column=col).value = value
    ws.row_dimensions[row_number].height = 75
    image_index = min(row_number - 4, 11) + 1
    image_path = THUMBS / f"product_{image_index:02d}.png"
    img = XLImage(str(image_path))
    img.width = 85
    img.height = 85
    ws.add_image(img, f"A{row_number}")

ws["A1"] = f"KATALOG DATA PRODUK & PROMO ALFAMIND ({len(rows)} ITEM)"
wb.save(WORKBOOK)
print(f"Updated {len(rows)} rows with product images")
