from pathlib import Path
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image

BASE = Path('D:/PROJECT/alfamind/ocr')
WORKBOOK = BASE / 'hasil_gantung11&12.xlsx'
THUMBS = BASE / 'gantung11_12_product_images'
THUMBS.mkdir(exist_ok=True)

def crop_gantung11():
    with Image.open(BASE / 'Gantung11.jpeg') as source:
        w, _ = source.size
        layouts = [(3, 140, 410), (3, 410, 670), (3, 670, 940), (1, 940, 1205)]
        idx = 1
        for count, y1, y2 in layouts:
            for col in range(count):
                x1, x2 = round(w * col / count), round(w * (col + 1) / count)
                source.crop((x1 + 8, y1 + 8, x2 - 8, y2 - 8)).save(THUMBS / f'g11_{idx:02d}.png')
                idx += 1

def crop_gantung12():
    with Image.open(BASE / 'Gantung12.jpeg') as source:
        w, _ = source.size
        ys = [140, 410, 670, 940, 1205]
        idx = 1
        for r in range(4):
            for c in range(3):
                x1, x2 = round(w * c / 3), round(w * (c + 1) / 3)
                source.crop((x1 + 8, ys[r] + 8, x2 - 8, ys[r + 1] - 8)).save(THUMBS / f'g12_{idx:02d}.png')
                idx += 1

crop_gantung11()
crop_gantung12()

wb = openpyxl.load_workbook(WORKBOOK)
ws = wb['Semua Produk']
mapping = [('g11', 1), ('g11', 2), ('g11', 2), ('g11', 3), ('g11', 3), ('g11', 3), ('g11', 4), ('g11', 5), ('g11', 6), ('g11', 7), ('g11', 8), ('g11', 9), ('g11', 10), ('g11', 10), ('g12', 1), ('g12', 1), ('g12', 1), ('g12', 1), ('g12', 2), ('g12', 3), ('g12', 4), ('g12', 5), ('g12', 5), ('g12', 5), ('g12', 6), ('g12', 6), ('g12', 7), ('g12', 7), ('g12', 7), ('g12', 8), ('g12', 9), ('g12', 9), ('g12', 10), ('g12', 11), ('g12', 12)]
for image in list(ws._images):
    ws._images.remove(image)

thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
white = PatternFill('solid', fgColor='FFFFFF')
zebra = PatternFill('solid', fgColor='F2F5F9')
stock_fill = PatternFill('solid', fgColor='E2EFDA')
stock_font = Font(name='Arial', size=10, color='375623', bold=True)

for row, (prefix, source_index) in enumerate(mapping, start=4):
    fill = zebra if (row - 4) % 2 else white
    ws.row_dimensions[row].height = 75
    for col in range(1, 16):
        cell = ws.cell(row, col)
        cell.font = Font(name='Arial', size=10)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=(col == 2))
    ws.cell(row, 6).number_format = '#,##0'
    ws.cell(row, 7).number_format = '#,##0'
    ws.cell(row, 6).alignment = Alignment(vertical='center', horizontal='right')
    ws.cell(row, 7).alignment = Alignment(vertical='center', horizontal='right')
    ws.cell(row, 9).fill = stock_fill
    ws.cell(row, 9).font = stock_font
    img = XLImage(str(THUMBS / f'{prefix}_{source_index:02d}.png'))
    img.width = 85
    img.height = 85
    ws.add_image(img, f'A{row}')

ws['A1'] = f'KATALOG DATA PRODUK & PROMO ALFAMIND ({len(mapping)} ITEM)'
ws.freeze_panes = 'A4'
wb.save(WORKBOOK)
print(f'Created {len(mapping)} rows with matching product images')
