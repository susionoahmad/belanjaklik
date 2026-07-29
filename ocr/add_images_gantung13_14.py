from pathlib import Path
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image

BASE = Path('D:/PROJECT/alfamind/ocr')
WORKBOOK = BASE / 'hasil_gantung13_14.xlsx'
THUMBS = BASE / 'gantung13_14_product_images'
THUMBS.mkdir(exist_ok=True)

def crop_grid(name, layouts, prefix):
    with Image.open(BASE / name) as source:
        w, _ = source.size
        idx = 1
        for count, y1, y2 in layouts:
            for col in range(count):
                x1, x2 = round(w * col / count), round(w * (col + 1) / count)
                source.crop((x1 + 8, y1 + 8, x2 - 8, y2 - 8)).save(THUMBS / f'{prefix}_{idx:02d}.png')
                idx += 1

crop_grid('Gantung13.jpeg', [(4, 140, 410), (4, 410, 670), (4, 670, 940), (4, 940, 1205)], 'g13')
crop_grid('Gantung14.jpeg', [(2, 140, 410), (4, 410, 670), (4, 670, 940), (4, 940, 1205)], 'g14')

wb = openpyxl.load_workbook(WORKBOOK)
ws = wb['Semua Produk']
mapping = []
for source, count in enumerate([3,3,1,2,3,2,2,2,2,3,2,1,4,2,2,1], start=1):
    mapping.extend([('g13', source)] * count)
for source, count in enumerate([1,1,1,1,1,1,1,1,1,1,1,1,1,1], start=1):
    mapping.extend([('g14', source)] * count)

for image in list(ws._images):
    ws._images.remove(image)

thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
white = PatternFill('solid', fgColor='FFFFFF')
zebra = PatternFill('solid', fgColor='F2F5F9')
stock_fill = PatternFill('solid', fgColor='E2EFDA')
stock_font = Font(name='Arial', size=10, color='375623', bold=True)

for row, (prefix, source_index) in enumerate(mapping, start=4):
    fill = zebra if (row - 4) % 2 else white
    ws.row_dimensions[row].height = 75
    for col in range(1, 16):
        cell = ws.cell(row, col)
        cell.font = Font(name='Arial', size=10)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=(col == 2))
    ws.cell(row, 6).number_format = '#,##0'
    ws.cell(row, 7).number_format = '#,##0'
    ws.cell(row, 6).alignment = Alignment(vertical='center', horizontal='right')
    ws.cell(row, 7).alignment = Alignment(vertical='center', horizontal='right')
    ws.cell(row, 9).fill = stock_fill
    ws.cell(row, 9).font = stock_font
    img = XLImage(str(THUMBS / f'{prefix}_{source_index:02d}.png'))
    img.width = 85
    img.height = 85
    ws.add_image(img, f'A{row}')

ws['A1'] = f'KATALOG DATA PRODUK & PROMO ALFAMIND ({len(mapping)} ITEM)'
ws.freeze_panes = 'A4'
wb.save(WORKBOOK)
print(f'Created {len(mapping)} rows with matching product images')
