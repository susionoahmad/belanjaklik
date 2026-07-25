import os
import sys
from pathlib import Path
from PIL import Image, ImageDraw

BASE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE_DIR))

# Force UTF-8 encoding for stdout on Windows
if sys.platform.startswith("win"):
    sys.stdout.reconfigure(encoding="utf-8")

from ocr_engine import extract_ocr_from_image
from excel_exporter import export_to_excel

def create_sample_promo_image(filepath: str):
    """Creates a sample JSM promo image for testing OCR pipeline."""
    img = Image.new('RGB', (800, 600), color=(255, 255, 255))
    draw = ImageDraw.Draw(img)

    # Header Promo JSM
    draw.rectangle([0, 0, 800, 80], fill=(220, 38, 38)) # Red Header
    draw.text((250, 20), "PROMO JSM ALFAMART", fill=(255, 255, 255))
    draw.text((260, 50), "PERIODE: 24 - 26 JULI 2026", fill=(255, 230, 0))

    # Product Card 1: Indomie
    draw.rectangle([30, 100, 370, 320], outline=(200, 200, 200), width=2)
    draw.text((50, 120), "INDOMIE GORENG SPESIAL 85G", fill=(0, 0, 0))
    draw.text((50, 150), "Rp 3.500", fill=(128, 128, 128)) # Strikethrough price
    draw.line([50, 158, 120, 158], fill=(220, 38, 38), width=2)
    draw.text((50, 180), "Rp 3.100", fill=(220, 38, 38)) # Discounted price
    draw.text((50, 210), "DISKON 11%", fill=(0, 150, 0))

    # Product Card 2: Bimoli
    draw.rectangle([430, 100, 770, 320], outline=(200, 200, 200), width=2)
    draw.text((450, 120), "BIMOLI MINYAK GORENG PCH 2L", fill=(0, 0, 0))
    draw.text((450, 150), "Rp 38.900", fill=(128, 128, 128))
    draw.line([450, 158, 530, 158], fill=(220, 38, 38), width=2)
    draw.text((450, 180), "Rp 34.900", fill=(220, 38, 38))
    draw.text((450, 210), "DISKON 10%", fill=(0, 150, 0))

    # Save Image
    img.save(filepath)
    print(f"[TEST] Sample promo image created at: {filepath}")

def run_demo():
    print("=" * 60)
    print(" DEMO RUNNER: ALFAMIND OCR & EXCEL EXPORTER")
    print("=" * 60)

    sample_img_path = BASE_DIR / "sample_promo_jsm.jpg"
    create_sample_promo_image(str(sample_img_path))

    output_excel_path = BASE_DIR / "hasil_ocr_import.xlsx"

    print("\n[INFO] Menjalankan Ekstraksi OCR pada sample promo...")
    products = extract_ocr_from_image(str(sample_img_path))

    print(f"\n[HASIL] Produk Terdeteksi: {len(products)}")
    for i, p in enumerate(products, 1):
        price_fmt = f"Rp{p.get('price'):,}" if isinstance(p.get('price'), (int, float)) and p.get('price') > 0 else "-"
        orig_fmt = f"Rp{p.get('original_price'):,}" if isinstance(p.get('original_price'), (int, float)) and p.get('original_price') > 0 else "-"
        print(f"  {i}. {p.get('product_name')} | {price_fmt} (Harga Asli: {orig_fmt}) | Promo: {p.get('promo_type')}")

    # Export to Excel
    export_to_excel(products, str(output_excel_path))
    
    print("\n[VERIFIKASI] Memeriksa header Excel vs kebutuhan Admin Import...")
    import openpyxl
    wb = openpyxl.load_workbook(output_excel_path)
    ws = wb.active
    header_row = [cell.value for cell in ws[1]]
    print(f"  Headers: {header_row}")
    print(f"  Total Data Rows: {ws.max_row - 1}")
    print("=" * 60)

if __name__ == "__main__":
    run_demo()
