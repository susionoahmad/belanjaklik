import os
import csv
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime

def export_to_excel(data_list: List[Dict[str, Any]], output_filepath: str, title_banner: str = "KATALOG DATA PRODUK & PROMO ALFAMIND") -> str:
    """
    Exports normalized product data to Excel (.xlsx) file with professional formatting:
    - Embedded thumbnail images (using OpenPyXLImage)
    - Styled Red Title Banner
    - Dark Blue Header Bar with White Bold Text
    - Zebra Striping (alternating row fills)
    - Stock Status Color Badges (Green for 'Tersedia', Red for 'Stok Kosong')
    - Currency Number Formatting (#,##0)
    - Freeze Panes & Auto Column Widths
    - Compatible with Admin Import (ExcelDriver.ts) schema
    """
    if not output_filepath.endswith(".xlsx") and not output_filepath.endswith(".csv"):
        output_filepath += ".xlsx"

    headers = [
        "image",
        "product_name",
        "brand",
        "variant",
        "package_size",
        "price",
        "original_price",
        "discount_percentage",
        "stock_status",
        "category",
        "promo_type",
        "promo_badge",
        "promo_title",
        "promo_start_date",
        "promo_end_date"
    ]

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.drawing.image import Image as OpenPyXLImage
        from PIL import Image as PILImage

        wb = openpyxl.Workbook()
        ws_master = wb.active
        ws_master.title = "Semua Produk"

        def build_sheet(ws, sheet_title_text, products):
            ws.views.sheetView[0].showGridLines = True

            # 1. Title Banner
            ws.merge_cells('A1:O1')
            ws['A1'] = sheet_title_text
            ws['A1'].font = Font(name='Arial', size=15, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid') # Alfamart Red
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 38

            ws['A2'] = "Ekstraksi Data Produk Ritel Lengkap (Katalog Master Toko Alfamind)"
            ws['A2'].font = Font(name='Arial', size=9, italic=True, color='555555')
            ws.row_dimensions[2].height = 18

            # 2. Table Headers (Row 3)
            header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid') # Navy Blue
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            ws.row_dimensions[3].height = 28

            # 3. Styling Tokens
            font_regular = Font(name='Arial', size=10)
            fill_zebra = PatternFill(start_color='F2F5F9', end_color='F2F5F9', fill_type='solid')
            fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

            fill_in_stock = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            font_in_stock = Font(name='Arial', size=10, color='375623', bold=True)

            fill_out_stock = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            font_out_stock = Font(name='Arial', size=10, color='C00000', bold=True)

            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )

            # Temp dir for generating cell thumbnail images
            temp_img_dir = Path(output_filepath).parent / "temp_thumbnails"
            temp_img_dir.mkdir(parents=True, exist_ok=True)

            start_row = 4
            for row_i, item in enumerate(products):
                row_num = start_row + row_i
                ws.row_dimensions[row_num].height = 75
                row_fill = fill_zebra if row_i % 2 == 1 else fill_white

                # Populate cells
                for col_idx, header_key in enumerate(headers, 1):
                    val = item.get(header_key, "")
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = font_regular
                    cell.border = thin_border
                    cell.fill = row_fill

                    if header_key in ["price", "original_price"]:
                        if val != "" and val is not None:
                            try:
                                cell.value = int(val)
                                cell.number_format = '#,##0'
                            except (ValueError, TypeError):
                                cell.value = val
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    elif header_key == "discount_percentage":
                        if val != "" and val is not None:
                            pct_str = str(val)
                            cell.value = f"{pct_str}%" if not pct_str.endswith('%') else pct_str
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif header_key == "stock_status":
                        status_str = str(val or "Tersedia").strip()
                        cell.value = status_str
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if "kosong" in status_str.lower() or "out" in status_str.lower():
                            cell.fill = fill_out_stock
                            cell.font = font_out_stock
                        else:
                            cell.fill = fill_in_stock
                            cell.font = font_in_stock
                    elif header_key in ["brand", "variant", "package_size", "category", "promo_type", "promo_badge", "promo_start_date", "promo_end_date"]:
                        cell.value = str(val or "")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif header_key == "product_name":
                        cell.value = str(val or "")
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    elif header_key == "image":
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        img_source = str(val or "").strip()
                        
                        # Embed thumbnail if image path exists
                        if img_source and os.path.exists(img_source):
                            try:
                                thumb_filename = temp_img_dir / f"thumb_{row_num}.png"
                                with PILImage.open(img_source) as p_img:
                                    canvas_thumb = PILImage.new('RGB', (90, 90), (255, 255, 255))
                                    crop_thumb = p_img.copy()
                                    crop_thumb.thumbnail((86, 86), PILImage.Resampling.LANCZOS)
                                    offset_t = ((90 - crop_thumb.width) // 2, (90 - crop_thumb.height) // 2)
                                    canvas_thumb.paste(crop_thumb, offset_t)
                                    canvas_thumb.save(thumb_filename)

                                img_obj = OpenPyXLImage(str(thumb_filename))
                                img_obj.width = 85
                                img_obj.height = 85
                                ws.add_image(img_obj, f'A{row_num}')
                            except Exception as img_err:
                                cell.value = img_source
                        else:
                            cell.value = img_source

            # Column Widths
            col_widths = {
                'A': 16, 'B': 38, 'C': 16, 'D': 24, 'E': 15,
                'F': 14, 'G': 15, 'H': 18, 'I': 16, 'J': 18,
                'K': 14, 'L': 22, 'M': 26, 'N': 16, 'O': 16
            }
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            ws.freeze_panes = 'A4'

        # Sheet 1: Master Sheet
        build_sheet(ws_master, f"{title_banner} ({len(data_list)} ITEM)", data_list)

        # Categorized Sheets (if multiple items)
        categories = set(p.get("category", "Umum") for p in data_list if p.get("category"))
        if len(categories) > 1 and len(data_list) > 3:
            for cat in sorted(categories):
                sheet_title = str(cat)[:30].replace("/", "-").replace("\\", "-")
                ws_cat = wb.create_sheet(title=sheet_title)
                cat_products = [p for p in data_list if p.get("category") == cat]
                build_sheet(ws_cat, f"KATALOG PRODUK {sheet_title.upper()} ({len(cat_products)} ITEM)", cat_products)

        try:
            wb.save(output_filepath)
            final_path = output_filepath
        except PermissionError:
            # File is open in Microsoft Excel, save with timestamp suffix
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            p = Path(output_filepath)
            final_path = str(p.parent / f"{p.stem}_{ts}{p.suffix}")
            print(f"[NOTICE] Target file '{output_filepath}' sedang dibuka di Microsoft Excel.")
            print(f"[INFO] Menyimpan hasil ke file alternatif: {final_path}")
            wb.save(final_path)

        print(f"[SUCCESS] Master Excel file created successfully: {final_path}")
        return final_path

    except ImportError:
        print("[WARNING] openpyxl/PIL not found. Falling back to CSV export...")
        csv_filepath = output_filepath.replace(".xlsx", ".csv")
        with open(csv_filepath, mode="w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for item in data_list:
                writer.writerow([item.get(h, "") for h in headers])
        print(f"[SUCCESS] CSV file saved to: {csv_filepath}")
        return csv_filepath
