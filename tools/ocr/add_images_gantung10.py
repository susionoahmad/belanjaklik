from pathlib import Path
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image

BASE = Path('D:/PROJECT/alfamind/ocr')
IMAGE = BASE / 'Gantung10.jpeg'
WORKBOOK = BASE / 'hasil_gantung10.xlsx'
THUMBS = BASE / 'gantung10_product_images'
THUMBS.mkdir(exist_ok=True)

with Image.open(IMAGE) as source:
    w, h = source.size
    # Card layout: 2 cards in row 1, then 4, 3, and 3 cards.
    layouts = [(2, 140, 410), (4, 410, 670), (3, 670, 940), (3, 940, 1205)]
    card_index = 1
    for count, y1, y2 in layouts:
        for col in range(count):
            x1, x2 = round(w * col / count), round(w * (col + 1) / count)
            source.crop((x1 + 8, y1 + 8, x2 - 8, y2 - 8)).save(THUMBS / f'product_{card_index:02d}.png')
            card_index += 1

wb = openpyxl.load_workbook(WORKBOOK)
ws = wb['Semua Produk']
source_indices = [1,1,1,1,2,3,3,3,4,5,6,7,7,7,8,9,10,11,11,12,12]
for image in list(ws._images):
    ws._images.remove(image)

thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
white = PatternFill('solid', fgColor='FFFFFF')
zebra = PatternFill('solid', fgColor='F2F5F9')
stock_fill = PatternFill('solid', fgColor='E2EFDA')
stock_font = Font(name='Arial', size=10, color='375623', bold=True)

for row_number, source_index in enumerate(source_indices, start=4):
    fill = zebra if (row_number - 4) % 2 else white
    ws.row_dimensions[row_number].height = 75
    for col in range(1, 16):
        cell = ws.cell(row_number, col)
        cell.font = Font(name='Arial', size=10)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=(col == 2))
    ws.cell(row_number, 6).number_format = '#,##0'
    ws.cell(row_number, 7).number_format = '#,##0'
    ws.cell(row_number, 6).alignment = Alignment(vertical='center', horizontal='right')
    ws.cell(row_number, 7).alignment = Alignment(vertical='center', horizontal='right')
    ws.cell(row_number, 9).fill = stock_fill
    ws.cell(row_number, 9).font = stock_font
    img = XLImage(str(THUMBS / f'product_{source_index:02d}.png'))
    img.width = 85
    img.height = 85
    ws.add_image(img, f'A{row_number}')

ws['A1'] = f'KATALOG DATA PRODUK & PROMO ALFAMIND ({len(source_indices)} ITEM)'
ws.freeze_panes = 'A4'
wb.save(WORKBOOK)
print(f'Created {len(source_indices)} rows with matching product images')
