from pathlib import Path
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image

BASE = Path('D:/PROJECT/alfamind/ocr')
IMAGE = BASE / 'Gantung9.jpeg'
WORKBOOK = BASE / 'hasil_gantung9.xlsx'
THUMBS = BASE / 'gantung9_product_images'
THUMBS.mkdir(exist_ok=True)

with Image.open(IMAGE) as source:
    w, h = source.size
    xs = [0, 240, 480, w]
    ys = [140, 410, 670, 940, 1205]
    for i in range(12):
        r, c = divmod(i, 3)
        source.crop((xs[c] + 8, ys[r] + 8, xs[c + 1] - 8, ys[r + 1] - 8)).save(THUMBS / f'product_{i + 1:02d}.png')

wb = openpyxl.load_workbook(WORKBOOK)
ws = wb['Semua Produk']
rows = []
for source_index, values in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True), start=1):
    name = str(values[1] or '')
    parts = [x.strip() for x in name.split(',') if x.strip()] or [name]
    rows.extend((values, part, source_index) for part in parts)

for image in list(ws._images):
    ws._images.remove(image)
ws.delete_rows(4, max(0, ws.max_row - 3))

thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
white = PatternFill('solid', fgColor='FFFFFF')
zebra = PatternFill('solid', fgColor='F2F5F9')
stock_fill = PatternFill('solid', fgColor='E2EFDA')
stock_font = Font(name='Arial', size=10, color='375623', bold=True)

for row_number, (values, name_part, source_index) in enumerate(rows, start=4):
    for col, value in enumerate(values, start=1):
        ws.cell(row=row_number, column=col).value = name_part if col == 2 else ('' if col == 1 else value)
    fill = zebra if (row_number - 4) % 2 else white
    ws.row_dimensions[row_number].height = 75
    for col in range(1, 16):
        cell = ws.cell(row_number, col)
        cell.font = Font(name='Arial', size=10)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=(col == 2))
    ws.cell(row_number, 6).number_format = '#,##0'
    ws.cell(row_number, 7).number_format = '#,##0'
    ws.cell(row_number, 6).alignment = Alignment(vertical='center', horizontal='right')
    ws.cell(row_number, 7).alignment = Alignment(vertical='center', horizontal='right')
    ws.cell(row_number, 9).fill = stock_fill
    ws.cell(row_number, 9).font = stock_font
    img = XLImage(str(THUMBS / f'product_{source_index:02d}.png'))
    img.width = 85
    img.height = 85
    ws.add_image(img, f'A{row_number}')

ws['A1'] = f'KATALOG DATA PRODUK & PROMO ALFAMIND ({len(rows)} ITEM)'
ws.freeze_panes = 'A4'
wb.save(WORKBOOK)
print(f'Created {len(rows)} rows with matching product images')
