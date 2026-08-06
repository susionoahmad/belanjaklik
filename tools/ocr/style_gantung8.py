from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

p = Path('D:/PROJECT/alfamind/ocr/hasil_gantung8.xlsx')
wb = openpyxl.load_workbook(p)
ws = wb['Semua Produk']
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
white = PatternFill('solid', fgColor='FFFFFF')
zebra = PatternFill('solid', fgColor='F2F5F9')
stock_fill = PatternFill('solid', fgColor='E2EFDA')
stock_font = Font(name='Arial', size=10, color='375623', bold=True)

for r in range(4, ws.max_row + 1):
    fill = zebra if (r - 4) % 2 else white
    ws.row_dimensions[r].height = 75
    for c in range(1, 16):
        cell = ws.cell(r, c)
        cell.font = Font(name='Arial', size=10)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=(c == 2))
    ws.cell(r, 6).number_format = '#,##0'
    ws.cell(r, 7).number_format = '#,##0'
    ws.cell(r, 6).alignment = Alignment(vertical='center', horizontal='right')
    ws.cell(r, 7).alignment = Alignment(vertical='center', horizontal='right')
    ws.cell(r, 9).fill = stock_fill
    ws.cell(r, 9).font = stock_font

ws.freeze_panes = 'A4'
ws.sheet_view.showGridLines = True
wb.save(p)
print(f'Formatted {ws.max_row - 3} data rows')
