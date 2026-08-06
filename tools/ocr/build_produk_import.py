from pathlib import Path
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "produk_import_engine.xlsx"
CROP_DIR = ROOT / "produk_import_crops"
CROP_DIR.mkdir(exist_ok=True)

products = [
    {"source": "produk1.jpeg", "crop": (100, 345, 315, 570), "product_name": "Bear Brand Collagen Vitamin C Susu Steril", "brand": "Bear Brand", "price": 11300},
    {"source": "produk1.jpeg", "crop": (390, 345, 620, 570), "product_name": "Collagena Susu Cair Kolagen Steril Kaleng", "brand": "Collagena", "price": 12500},
    {"source": "produk1.jpeg", "crop": (100, 1015, 315, 1240), "product_name": "Nestle Goodnes Kurma Ajwa Madinah Susu Steril", "brand": "Nestle Goodnes", "price": 10500},
    {"source": "produk1.jpeg", "crop": (390, 1015, 620, 1240), "product_name": "Nestle Goodnes Kurma & Madu Habbatussauda", "brand": "Nestle Goodnes", "price": 10500},
    {"source": "produk2.jpeg", "crop": (100, 345, 315, 570), "product_name": "Bear Brand Susu Steril Kaleng 189 ml", "brand": "Bear Brand", "price": 10800},
    {"source": "produk2.jpeg", "crop": (390, 345, 620, 570), "product_name": "So Good Susu Cair Steril Kaleng 189 ml", "brand": "So Good", "price": 8900},
    {"source": "produk2.jpeg", "crop": (100, 1015, 315, 1240), "product_name": "Entrasol O'live Milk Susu Cair Steril Ekstrak Buah", "brand": "Entrasol", "price": 10000},
    {"source": "produk2.jpeg", "crop": (390, 1015, 620, 1240), "product_name": "Bear Brand Gold Susu Steril Malt Kaleng 140 ml", "brand": "Bear Brand", "price": 11900},
]

headers = [
    "image", "product_name", "brand", "variant", "package_size", "price", "original_price",
    "discount_percentage", "stock_status", "category", "promo_type", "promo_badge", "promo_title",
    "promo_start_date", "promo_end_date"
]

wb = Workbook()
ws = wb.active
ws.title = "Product Import"
ws.freeze_panes = "A2"
ws.sheet_view.showGridLines = False
ws.append(headers)

for idx, product in enumerate(products, start=2):
    ws.append([
        "", product["product_name"], product["brand"], "", "", product["price"], "", "",
        "Tersedia", "Susu Cair", "REGULAR", "", "", "", ""
    ])
    ws.row_dimensions[idx].height = 92
    source = Image.open(ROOT / product["source"]).convert("RGB")
    crop = source.crop(product["crop"])
    crop.thumbnail((150, 120), Image.Resampling.LANCZOS)
    crop_path = CROP_DIR / f"product_{idx-1}.png"
    crop.save(crop_path, "PNG")
    img = XLImage(str(crop_path))
    img.width = crop.width
    img.height = crop.height
    ws.add_image(img, f"A{idx}")

# Readable import layout.
widths = {"A": 20, "B": 48, "C": 18, "D": 18, "E": 16, "F": 14, "G": 16, "H": 18, "I": 16, "J": 16, "K": 14, "L": 18, "M": 34, "N": 16, "O": 16}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

header_fill = PatternFill("solid", fgColor="C00000")
header_font = Font(color="FFFFFF", bold=True)
thin_gray = Side(style="thin", color="D9D9D9")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=thin_gray)
ws.row_dimensions[1].height = 30

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)
    row[5].number_format = 'Rp #,##0'
    row[6].number_format = 'Rp #,##0'

ref = f"A1:O{ws.max_row}"
tab = Table(displayName="ProductImportTable", ref=ref)
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
ws.add_table(tab)

wb.save(OUT)
print(OUT)
